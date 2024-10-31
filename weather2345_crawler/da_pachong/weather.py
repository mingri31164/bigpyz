import asyncio
import aiohttp
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

class crawler_meta:
    def __init__(self, browser):
        self.browser = browser

    def drop_scroll(self):
        '''滚动页面到最后'''
        last_height = self.browser.execute_script("return document.documentElement.scrollHeight")
        while True:
            self.browser.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
            time.sleep(2)
            new_height = self.browser.execute_script("return document.documentElement.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    def start_crawl(self):
        '''开始爬取'''
        raise NotImplementedError("请实现执行方法")
    
    def get_page_data(self):
        pass

class tenkiCrawler(crawler_meta):
    def __init__(self, tenki_url_raw: str, relative_path='my_tenki_infos'):
        self.tenki_url_raw = tenki_url_raw
        self.relative_path = relative_path
        self.browser = None
        self.weather_data = []  # 存储处理后的天气数据

    def process_weather_data(self, items_text: List[str]) -> None:
        '''处理天气数据'''
        for day_data in items_text:
            if day_data and len(day_data.strip()) > 0:
                # 跳过表头
                if '日期' in day_data or len(day_data.strip()) < 10:  # 增加最小长度判断
                    continue
                
                try:
                    # 使用空格分割数据
                    parts = day_data.strip().split()
                    
                    # 确保至少有基本的数据字段
                    if len(parts) >= 5:
                        date = parts[0]  # 日期
                        # 去掉温度中的度数符号
                        high_temp = parts[2].replace('°', '').strip()
                        low_temp = parts[3].replace('°', '').strip()
                        
                        # 提取天气状况（可能包含~符号的组合天气）
                        weather = parts[4]
                        
                        # 添加到数据列表
                        self.weather_data.append({
                            'date': date,
                            'high_temp': high_temp,
                            'low_temp': low_temp,
                            'weather': weather
                        })
                except Exception as e:
                    print(f"处理数据时出错: {e}")
                    print(f"问题数据: {day_data}")
                    continue

    def save_to_excel(self, filename: str = None) -> None:
        '''保存数据到Excel文件'''
        if filename is None:
            # 使用当前时间创建文件名
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'weather_data_{current_time}.xlsx'

        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "天气数据"

            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 日期
            ws.column_dimensions['B'].width = 12  # 最高温度
            ws.column_dimensions['C'].width = 12  # 最低温度
            ws.column_dimensions['D'].width = 20  # 天气状况

            # 写入表头
            headers = ['日期', '最高温度(℃)', '最低温度(℃)', '天气状况']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # 写入数据
            for row, data in enumerate(self.weather_data, 2):
                # 日期直接写入，不需要转换
                ws.cell(row=row, column=1, value=data['date'])
                
                try:
                    # 温度转换为数字，如果转换失败则保持原样
                    high_temp = float(data['high_temp']) if data['high_temp'].replace('-', '').isdigit() else data['high_temp']
                    low_temp = float(data['low_temp']) if data['low_temp'].replace('-', '').isdigit() else data['low_temp']
                    
                    ws.cell(row=row, column=2, value=high_temp)
                    ws.cell(row=row, column=3, value=low_temp)
                except ValueError:
                    # 如果转换失败，直接写入原始字符串
                    ws.cell(row=row, column=2, value=data['high_temp'])
                    ws.cell(row=row, column=3, value=data['low_temp'])
                
                ws.cell(row=row, column=4, value=data['weather'])
                
                # 为数据单元格添加样式
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = border

            # 保存文件
            wb.save(filename)
            print(f"数据已成功保存到 {filename}")
        except Exception as e:
            print(f"保存Excel文件时出错: {e}")
            raise  # 抛出异常以便调试

    def get_page_data(self):
        '''页面数据获取'''
        tenki_items = []
        tenki_items_text = []
        
        for i in range(20):  # 获取20个月的数据
            time.sleep(2)
            wait = WebDriverWait(self.browser, 10)

            tenki_items_ = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//tr")))
            tenki_items.append(tenki_items_)
            tenki_items_text_ = [item.text for item in tenki_items_]
            tenki_items_text.append(tenki_items_text_)
            
            # 处理当前页面的数据
            self.process_weather_data(tenki_items_text_)
            
            try:
                node = self.browser.find_element(By.XPATH, "//a[contains(@id, 'js_prevMonth')]")
                node.click()
            except Exception as e:
                print(f"翻页时出错: {e}")
                break

    def start_crawl(self):
        '''开始爬取'''
        options = Options()
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0")
        self.browser = webdriver.Chrome(options=options)

        try:
            self.browser.get(self.tenki_url_raw)
            time.sleep(2)  # 等待页面加载
            
            self.get_page_data()
            
            # 保存数据到Excel
            self.save_to_excel()
           
        finally:
            # 关闭浏览器
            self.browser.quit()

        print("爬取完成，数据已保存。")

if __name__ == '__main__':
    # 开始爬取
    crawler = tenkiCrawler(
        tenki_url_raw='https://tianqi.2345.com/wea_history/59431.htm',
        relative_path='my_tenki_infos'
    )
    crawler.start_crawl()